import openpyxl

# Cargar el archivo de Excel
workbook = openpyxl.load_workbook("asistencia.xlsx")
sheet = workbook["Asistencia"]

# Obtener la última fila con datos
last_row = sheet.max_row

# Agregar una nueva fila
sheet.insert_rows(last_row + 1)

# Solicitar datos al usuario
nombre = input("Ingrese su nombre: ")
fecha = input("Ingrese la fecha (AAAA-MM-DD): ")
hora = input("Ingrese la hora de entrada (HH:MM): ")

# Escribir los datos en la nueva fila
sheet.cell(row=last_row + 1, column=1).value = nombre
sheet.cell(row=last_row + 1, column=2).value = fecha
sheet.cell(row=last_row + 1, column=3).value = hora

# Guardar los cambios en el archivo
workbook.save("asistencia.xlsx")